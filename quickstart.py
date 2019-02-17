#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
hows basic usage of the Google Calendar API. Creates a Google Calendar API
service object and outputs a list of the next 10 events on the user's calendar.
"""
from __future__ import print_function
from apiclient.discovery import build
from httplib2 import Http
from oauth2client import file, client, tools
import datetime
import sys 
import argparse

#sys.setdefaultencoding('utf-8')
from openpyxl import load_workbook
from openpyxl import Workbook


def parsArgs():
    parser = argparse.ArgumentParser()
    parser.add_argument("--email", help="users google acount mail", action="store")
    parser.add_argument("--name", help="users name as shown in exle file", action="store")
    parser.add_argument("--xls", help="xls toranut fike", action="store")
    #parser.add_argument("--dates", nargs='+' , help="days of toranut", action="store")
    args = parser.parse_args()
    return args


my_email='neiterman21@gmail.com'
global service


# Setup the Calendar API
def setup_calendar():
    global service
    SCOPES = 'https://www.googleapis.com/auth/calendar'
    store = file.Storage('credentials.json')
    creds = store.get()
    if not creds or creds.invalid:
        flow = client.flow_from_clientsecrets('client_secret.json', SCOPES)
        creds = tools.run_flow(flow, store)
    service = build('calendar', 'v3', http=creds.authorize(Http()))
def pars_xls(xls_file):
    return 0

def pars_dates(dates):
    ret_dates = []
    for date in dates:
        if (isinstance(date, datetime.datetime)):
            #the mox betwing day and mounth is due to date missuse in the xls file. NOT A BUG
            ret_dates.append(str(date.month)  + "/" + str(date.day) + "/" + str(date.year)[-2:])

        if (isinstance(date, unicode)):
            ret_dates.append(date.encode('utf8'))
    return ret_dates

def toranut_dates(name , xls_string):
    wb = load_workbook(xls_string)
    ws = wb.active
    dates = []
    for x in range(2,33):
        if name == ws.cell(row=x, column=4).value or name == ws.cell(row=x, column=5).value or name == ws.cell(row=x, column=6).value or name == ws.cell(row=x, column=7).value:
            dates.append(ws.cell(row=x, column=3).value)
    return pars_dates(dates)


# Call the Calendar API
def Main():
    global service
    args = parsArgs()
    setup_calendar()
#    toranuyot = pars_xls(args.xls)
    colors = service.colors().get().execute()
    myevent = {
    'summary': 'תורנות',
    'location': 'הילל יפה',
    'description': 'הזדמנות לא לישון מחוץ לבית',
    'start': {
        'date': '2018-10-22'
    },
    'end': {
        'date': '2018-10-22'
    },
    'recurrence': [
        'RRULE:FREQ=DAILY;COUNT=1'
    ],
    'attendees': [],
    'colorId' : "11",
    'reminders': {
        'useDefault': False,
        'overrides': [],
    },
    }
    date = '%s-%s-%s'
    dates = toranut_dates(unicode(args.name, 'utf8') , args.xls)
    for day in dates :
        split_day = day.split('/')
        myevent['start']['date'] = ( "20" + split_day[2] + "-" + split_day[1] + "-" + split_day[0])
        myevent['end']['date']   = myevent['start']['date']
        print (myevent['start']['date'])
        service.events().insert(calendarId=args.email , body = myevent).execute()
    '''
    for day in toranuyot:
        if args.name in day['toranim']:
            myevent['start']['date'] = day['date']
            myevent['end']['date'] = day['date']
            myevent['colorId'] = 'Tomato'
            service.events().insert(calendarId=args.email , body = myevent).execute()
        if args.name in day['vacation']:
            myevent['start']['date'] = day['date']
            myevent['end']['date'] = day['date']
            myevent['colorId'] = 'Tomato'
            service.events().insert(calendarId=args.email , body = myevent).execute()
    '''

if __name__ == '__main__':
    Main()
